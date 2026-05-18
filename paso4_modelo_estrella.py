import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
#  PASO 4 — Modelo de Datos Estrella
#  Construye 5 tablas:
#    · Hechos_Ventas  (tabla de hechos)
#    · Dim_Tiempo     (dimensión tiempo)
#    · Dim_Cliente    (dimensión cliente)
#    · Dim_Producto   (dimensión producto)
#    · Dim_Region     (dimensión región geográfica)
#  Exporta todo a un Excel con diseño profesional.
# ─────────────────────────────────────────────

# ── Paleta de colores por tabla ───────────────
COLORES = {
    "Hechos_Ventas": {"header": "C0392B", "subheader": "E74C3C"},  # Rojo
    "Dim_Tiempo":    {"header": "1A5276", "subheader": "2980B9"},  # Azul
    "Dim_Cliente":   {"header": "1E8449", "subheader": "27AE60"},  # Verde
    "Dim_Producto":  {"header": "6C3483", "subheader": "8E44AD"},  # Morado
    "Dim_Region":    {"header": "784212", "subheader": "E67E22"},  # Naranja
}

MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo",  6: "Junio",  7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}

print("=" * 60)
print("  PASO 4 — MODELO DE DATOS ESTRELLA")
print("=" * 60)

# ── Cargar datos del paso anterior ────────────
print("\nCargando datos del Paso 3...")
df = pd.read_pickle("datos_paso3.pkl")
print(f"  Registros: {len(df):,}  |  Columnas: {len(df.columns)}")


# ══════════════════════════════════════════════
#  1. DIM_TIEMPO
# ══════════════════════════════════════════════
print("\n" + "─" * 60)
print("Construyendo Dim_Tiempo...")

todas_fechas = pd.concat([df["Fecha_Salida"], df["Fecha_Entrega"]]).drop_duplicates()
dim_tiempo = pd.DataFrame({"Fecha": todas_fechas.sort_values()})
dim_tiempo["ID_Tiempo"]   = dim_tiempo["Fecha"].dt.strftime("%Y%m%d").astype(int)
dim_tiempo["Año"]         = dim_tiempo["Fecha"].dt.year
dim_tiempo["Trimestre"]   = dim_tiempo["Fecha"].dt.quarter.map(
                                lambda q: f"T{q}")
dim_tiempo["Mes"]         = dim_tiempo["Fecha"].dt.month
dim_tiempo["Nombre_Mes"]  = dim_tiempo["Mes"].map(MESES_ES)
dim_tiempo["Dia"]         = dim_tiempo["Fecha"].dt.day
dim_tiempo["Dia_Semana"]  = dim_tiempo["Fecha"].dt.day_name()
dim_tiempo = dim_tiempo[["ID_Tiempo", "Fecha", "Año", "Trimestre",
                          "Mes", "Nombre_Mes", "Dia", "Dia_Semana"]]
dim_tiempo = dim_tiempo.reset_index(drop=True)
print(f"  Filas generadas: {len(dim_tiempo):,}  |  Columnas: {len(dim_tiempo.columns)}")

# Mapa fecha → ID_Tiempo para la tabla de hechos
mapa_tiempo = dim_tiempo.set_index("Fecha")["ID_Tiempo"].to_dict()


# ══════════════════════════════════════════════
#  2. DIM_CLIENTE
# ══════════════════════════════════════════════
print("\n" + "─" * 60)
print("Construyendo Dim_Cliente...")

dim_cliente = (df[["ID_Cliente", "Número Cliente", "Nombre Cliente", "Segmento"]]
               .drop_duplicates(subset="ID_Cliente")
               .sort_values("ID_Cliente")
               .reset_index(drop=True))
dim_cliente.columns = ["ID_Cliente", "Número_Cliente", "Nombre_Cliente", "Segmento"]
print(f"  Filas generadas: {len(dim_cliente):,}  |  Columnas: {len(dim_cliente.columns)}")


# ══════════════════════════════════════════════
#  3. DIM_PRODUCTO
# ══════════════════════════════════════════════
print("\n" + "─" * 60)
print("Construyendo Dim_Producto...")

dim_producto = (df[["ID Producto", "Categoría", "Sub-Categoría", "Nombre Producto"]]
                .drop_duplicates(subset="ID Producto")
                .sort_values("ID Producto")
                .reset_index(drop=True))
dim_producto.columns = ["ID_Producto", "Categoría", "Sub_Categoría", "Nombre_Producto"]
print(f"  Filas generadas: {len(dim_producto):,}  |  Columnas: {len(dim_producto.columns)}")


# ══════════════════════════════════════════════
#  4. DIM_REGION
# ══════════════════════════════════════════════
print("\n" + "─" * 60)
print("Construyendo Dim_Region...")

dim_region = (df[["Ciudad", "Estado", "País", "Mercado", "Región"]]
              .drop_duplicates()
              .sort_values(["País", "Estado", "Ciudad"])
              .reset_index(drop=True))
dim_region.insert(0, "ID_Region", range(1, len(dim_region) + 1))
print(f"  Filas generadas: {len(dim_region):,}  |  Columnas: {len(dim_region.columns)}")

# Mapa (Ciudad, Estado, País) → ID_Region para la tabla de hechos
mapa_region = dim_region.set_index(["Ciudad", "Estado", "País"])["ID_Region"].to_dict()


# ══════════════════════════════════════════════
#  5. HECHOS_VENTAS
# ══════════════════════════════════════════════
print("\n" + "─" * 60)
print("Construyendo Hechos_Ventas...")

hechos = df.copy()

# Claves foráneas
hechos["FK_Tiempo_Salida"]  = hechos["Fecha_Salida"].map(mapa_tiempo)
hechos["FK_Tiempo_Entrega"] = hechos["Fecha_Entrega"].map(mapa_tiempo)
hechos["FK_Region"]         = hechos.apply(
    lambda r: mapa_region.get((r["Ciudad"], r["Estado"], r["País"]), None), axis=1)

# Clave primaria secuencial
hechos.insert(0, "ID_Hecho", range(1, len(hechos) + 1))

# Seleccionar solo las columnas necesarias (orden lógico)
hechos_ventas = hechos[[
    "ID_Hecho",
    "ID_Venta",
    "Número Venta",
    "FK_Tiempo_Salida",
    "FK_Tiempo_Entrega",
    "ID_Cliente",
    "ID Producto",
    "FK_Region",
    "Método Envio",
    "Prioridad Envio",
    "Ventas",
    "Cantidad",
    "Descuento",
    "Utilidad",
    "Costo Envío",
]].copy()

hechos_ventas.columns = [
    "ID_Hecho",
    "ID_Venta",
    "Número_Venta",
    "FK_Tiempo_Salida",
    "FK_Tiempo_Entrega",
    "FK_Cliente",
    "FK_Producto",
    "FK_Region",
    "Método_Envio",
    "Prioridad_Envio",
    "Ventas",
    "Cantidad",
    "Descuento",
    "Utilidad",
    "Costo_Envio",
]

print(f"  Filas generadas: {len(hechos_ventas):,}  |  Columnas: {len(hechos_ventas.columns)}")

# Verificar integridad referencial
nulos_fk = {
    "FK_Tiempo_Salida":  hechos_ventas["FK_Tiempo_Salida"].isna().sum(),
    "FK_Tiempo_Entrega": hechos_ventas["FK_Tiempo_Entrega"].isna().sum(),
    "FK_Cliente":        hechos_ventas["FK_Cliente"].isna().sum(),
    "FK_Producto":       hechos_ventas["FK_Producto"].isna().sum(),
    "FK_Region":         hechos_ventas["FK_Region"].isna().sum(),
}
print("\n  Verificación de integridad referencial (FK nulos):")
for fk, val in nulos_fk.items():
    estado = "OK" if val == 0 else f"AVISO: {val} nulos"
    print(f"    {fk}: {estado}")


# ══════════════════════════════════════════════
#  6. EXPORTAR A EXCEL CON DISEÑO
# ══════════════════════════════════════════════
print("\n" + "─" * 60)
print("Exportando Modelo Estrella a Excel...")

NOMBRE_ARCHIVO = "Modelo_Estrella_Ventas.xlsx"

tablas = [
    ("Hechos_Ventas", hechos_ventas),
    ("Dim_Tiempo",    dim_tiempo),
    ("Dim_Cliente",   dim_cliente),
    ("Dim_Producto",  dim_producto),
    ("Dim_Region",    dim_region),
]

def borde_fino():
    lado = Side(style="thin", color="BDBDBD")
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def borde_medio():
    lado = Side(style="medium", color="666666")
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def aplicar_hoja(wb, nombre_hoja, df_tabla, color_header, color_sub):
    ws = wb.create_sheet(title=nombre_hoja)

    fill_header = PatternFill("solid", fgColor=color_header)
    fill_sub    = PatternFill("solid", fgColor=color_sub)
    fill_par    = PatternFill("solid", fgColor="F8F9FA")   # filas pares
    fill_impar  = PatternFill("solid", fgColor="FFFFFF")   # filas impares
    fuente_header = Font(bold=True, color="FFFFFF", size=11)
    fuente_datos  = Font(size=10)

    cols = list(df_tabla.columns)

    # ── Fila 1: título de la hoja ──
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=len(cols))
    celda_titulo = ws.cell(row=1, column=1, value=f"  {nombre_hoja}")
    celda_titulo.fill      = fill_header
    celda_titulo.font      = Font(bold=True, color="FFFFFF", size=13)
    celda_titulo.alignment = Alignment(horizontal="left", vertical="center",
                                       indent=1)
    ws.row_dimensions[1].height = 28

    # ── Fila 2: encabezados de columnas ──
    for c_idx, col in enumerate(cols, start=1):
        celda = ws.cell(row=2, column=c_idx, value=col)
        celda.fill      = fill_sub
        celda.font      = fuente_header
        celda.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
        celda.border    = borde_medio()
    ws.row_dimensions[2].height = 22

    # ── Filas de datos ──
    for r_idx, fila in enumerate(df_tabla.itertuples(index=False), start=3):
        fill_fila = fill_par if r_idx % 2 == 0 else fill_impar
        for c_idx, valor in enumerate(fila, start=1):
            # Formatear fechas
            if isinstance(valor, pd.Timestamp):
                valor = valor.strftime("%Y-%m-%d")
            celda = ws.cell(row=r_idx, column=c_idx, value=valor)
            celda.fill      = fill_fila
            celda.font      = fuente_datos
            celda.border    = borde_fino()
            # Alinear números a la derecha
            col_nombre = cols[c_idx - 1]
            if df_tabla[col_nombre].dtype in ["int64", "float64"]:
                celda.alignment = Alignment(horizontal="right",
                                            vertical="center")
                if "Ventas" in col_nombre or "Utilidad" in col_nombre \
                        or "Costo" in col_nombre:
                    celda.number_format = '#,##0.00'
                elif "Descuento" in col_nombre:
                    celda.number_format = '0.00%'
            else:
                celda.alignment = Alignment(horizontal="left",
                                            vertical="center")

    # ── Ajustar ancho de columnas ──
    for c_idx, col in enumerate(cols, start=1):
        letra = get_column_letter(c_idx)
        max_len = max(
            len(str(col)),
            df_tabla[col].astype(str).str.len().max() if len(df_tabla) > 0 else 0,
        )
        ws.column_dimensions[letra].width = min(max_len + 4, 45)

    # ── Inmovilizar fila de encabezados ──
    ws.freeze_panes = "A3"

    print(f"  Hoja '{nombre_hoja}' — {len(df_tabla):,} filas  x  {len(cols)} columnas")
    return ws


wb = Workbook()
wb.remove(wb.active)  # eliminar hoja por defecto

for nombre, tabla in tablas:
    colores = COLORES[nombre]
    aplicar_hoja(wb, nombre, tabla,
                 colores["header"], colores["subheader"])

wb.save(NOMBRE_ARCHIVO)
print(f"\n  Archivo guardado: {NOMBRE_ARCHIVO}")


# ── Guardar tablas para el siguiente paso ─────
print("\nGuardando tablas para uso en siguientes pasos...")
hechos_ventas.to_pickle("hechos_ventas.pkl")
dim_tiempo.to_pickle("dim_tiempo.pkl")
dim_cliente.to_pickle("dim_cliente.pkl")
dim_producto.to_pickle("dim_producto.pkl")
dim_region.to_pickle("dim_region.pkl")
print("  Archivos guardados correctamente.")


# ── Resumen final ─────────────────────────────
print("\n" + "=" * 60)
print("  RESUMEN — Modelo Estrella construido")
print("=" * 60)
resumen = [
    ("Hechos_Ventas", len(hechos_ventas), len(hechos_ventas.columns),
     "Tabla central con métricas"),
    ("Dim_Tiempo",    len(dim_tiempo),    len(dim_tiempo.columns),
     "Fechas de salida y entrega"),
    ("Dim_Cliente",   len(dim_cliente),   len(dim_cliente.columns),
     "Datos de clientes"),
    ("Dim_Producto",  len(dim_producto),  len(dim_producto.columns),
     "Catálogo de productos"),
    ("Dim_Region",    len(dim_region),    len(dim_region.columns),
     "Geografía: ciudad/país/mercado"),
]
print(f"\n  {'Tabla':<20} {'Filas':>8} {'Cols':>5}  Descripción")
print("  " + "─" * 58)
for nombre, filas, cols, desc in resumen:
    print(f"  {nombre:<20} {filas:>8,} {cols:>5}  {desc}")

print("\n¡Paso 4 completado exitosamente!")
