import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
#  PASO 6 — Tablas Dinámicas
#  Ejercicio 1: Ventas por Mercado
#  Ejercicio 2: Ventas por Año y Mes
#  Ejercicio 3: % de Ventas (vs total) por Método de Envío
#  Ejercicio 4: Top 3 Mercados con más Ventas por Prioridad de Envío
# ─────────────────────────────────────────────

MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo",    4: "Abril",
    5: "Mayo",  6: "Junio",  7: "Julio",     8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}

ORDEN_PRIORIDAD = ["Critical", "High", "Medium", "Low"]

print("=" * 65)
print("  PASO 6 — TABLAS DINÁMICAS")
print("=" * 65)

# ── Cargar datos ──────────────────────────────
print("\nCargando datos del Paso 3...")
df = pd.read_pickle("datos_paso3.pkl")
ventas_total = df["Ventas"].sum()
print(f"  Registros: {len(df):,}  |  Ventas totales: ${ventas_total:,.2f}")


# ══════════════════════════════════════════════
#  EJERCICIO 1 — Ventas por Mercado
# ══════════════════════════════════════════════
print("\n" + "─" * 65)
print("EJERCICIO 1 — Ventas por Mercado")
print("─" * 65)

t1 = (df.groupby("Mercado", as_index=False)["Ventas"]
        .sum()
        .sort_values("Ventas", ascending=False)
        .reset_index(drop=True))
t1.index += 1
t1["% del Total"] = (t1["Ventas"] / ventas_total * 100).round(2)
t1["Ventas"]      = t1["Ventas"].round(2)

print(f"\n{'#':>3}  {'Mercado':<10}  {'Ventas ($)':>15}  {'% del Total':>12}")
print("  " + "─" * 47)
for idx, row in t1.iterrows():
    print(f"  {idx:>2}.  {row['Mercado']:<10}  "
          f"${row['Ventas']:>14,.2f}  {row['% del Total']:>11.2f}%")
print(f"\n  Total: ${ventas_total:>,.2f}")


# ══════════════════════════════════════════════
#  EJERCICIO 2 — Ventas por Año y Mes
# ══════════════════════════════════════════════
print("\n" + "─" * 65)
print("EJERCICIO 2 — Ventas por Año y Mes")
print("─" * 65)

t2 = (df.groupby(["Año Salida", "Mes Salida"], as_index=False)["Ventas"]
        .sum()
        .sort_values(["Año Salida", "Mes Salida"])
        .reset_index(drop=True))
t2["Nombre_Mes"] = t2["Mes Salida"].map(MESES_ES)
t2["Ventas"]     = t2["Ventas"].round(2)
t2 = t2.rename(columns={"Año Salida": "Año", "Mes Salida": "Mes"})
t2 = t2[["Año", "Mes", "Nombre_Mes", "Ventas"]]

print(f"\n  {'Año':>4}  {'Mes':>3}  {'Nombre Mes':<13}  {'Ventas ($)':>15}")
print("  " + "─" * 42)
año_actual = None
for _, row in t2.iterrows():
    if año_actual != row["Año"]:
        if año_actual is not None:
            # Subtotal por año
            sub = t2[t2["Año"] == año_actual]["Ventas"].sum()
            print(f"  {'':>4}  {'':>3}  {'  SUBTOTAL ' + str(año_actual):<13}  "
                  f"${sub:>14,.2f}")
            print("  " + "─" * 42)
        año_actual = row["Año"]
    print(f"  {int(row['Año']):>4}  {int(row['Mes']):>3}  "
          f"{row['Nombre_Mes']:<13}  ${row['Ventas']:>14,.2f}")
# Último subtotal
sub = t2[t2["Año"] == año_actual]["Ventas"].sum()
print(f"  {'':>4}  {'':>3}  {'  SUBTOTAL ' + str(año_actual):<13}  "
      f"${sub:>14,.2f}")


# ══════════════════════════════════════════════
#  EJERCICIO 3 — % de Ventas por Método de Envío
# ══════════════════════════════════════════════
print("\n" + "─" * 65)
print("EJERCICIO 3 — % de Ventas (vs total) por Método de Envío")
print("─" * 65)

t3 = (df.groupby("Método Envio", as_index=False)["Ventas"]
        .sum()
        .sort_values("Ventas", ascending=False)
        .reset_index(drop=True))
t3.index += 1
t3["% del Total"]    = (t3["Ventas"] / ventas_total * 100).round(2)
t3["Ventas"]         = t3["Ventas"].round(2)
t3 = t3.rename(columns={"Método Envio": "Método de Envío"})

print(f"\n  {'#':>2}  {'Método de Envío':<17}  {'Ventas ($)':>15}  {'% del Total':>12}")
print("  " + "─" * 52)
for idx, row in t3.iterrows():
    barra = "█" * int(row["% del Total"] / 2)
    print(f"  {idx:>2}.  {row['Método de Envío']:<17}  "
          f"${row['Ventas']:>14,.2f}  {row['% del Total']:>10.2f}%  {barra}")
print(f"\n  Total: ${ventas_total:>,.2f}  (100.00%)")


# ══════════════════════════════════════════════
#  EJERCICIO 4 — Top 3 Mercados por Prioridad de Envío
# ══════════════════════════════════════════════
print("\n" + "─" * 65)
print("EJERCICIO 4 — Top 3 Mercados con más Ventas por Prioridad de Envío")
print("─" * 65)

t4_full = (df.groupby(["Prioridad Envio", "Mercado"], as_index=False)["Ventas"]
             .sum())

# Top 3 por prioridad — construido fila a fila para evitar problemas de versión
registros_t4 = []
for prioridad in ORDEN_PRIORIDAD:
    grupo = (t4_full[t4_full["Prioridad Envio"] == prioridad]
             .nlargest(3, "Ventas"))
    total_prior = t4_full[t4_full["Prioridad Envio"] == prioridad]["Ventas"].sum()
    for puesto, (_, row) in enumerate(grupo.iterrows(), start=1):
        registros_t4.append({
            "Prioridad de Envío":  prioridad,
            "Puesto":              puesto,
            "Mercado":             row["Mercado"],
            "Ventas":              round(row["Ventas"], 2),
            "% de su Prioridad":   round(row["Ventas"] / total_prior * 100, 2),
        })

t4 = pd.DataFrame(registros_t4)

print()
prioridad_actual = None
for _, row in t4.iterrows():
    if prioridad_actual != row["Prioridad de Envío"]:
        prioridad_actual = row["Prioridad de Envío"]
        total_prior = t4_full[t4_full["Prioridad Envio"] == prioridad_actual
                               ]["Ventas"].sum()
        print(f"\n  Prioridad: {prioridad_actual}  "
              f"(Ventas totales: ${total_prior:,.2f})")
        print(f"  {'Puesto':>6}  {'Mercado':<10}  {'Ventas ($)':>15}  "
              f"{'% Prioridad':>13}")
        print("  " + "─" * 50)
    medalla = ["🥇", "🥈", "🥉"][row["Puesto"] - 1]
    print(f"  {medalla} {row['Puesto']:>4}°  {row['Mercado']:<10}  "
          f"${row['Ventas']:>14,.2f}  {row['% de su Prioridad']:>12.2f}%")


# ══════════════════════════════════════════════
#  EXPORTAR A EXCEL CON DISEÑO
# ══════════════════════════════════════════════
print("\n\n" + "─" * 65)
print("Exportando Tablas Dinámicas a Excel...")

NOMBRE_ARCHIVO = "Tablas_Dinamicas_Ventas.xlsx"

# ── helpers de estilo ─────────────────────────
def borde_fino():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def borde_medio():
    s = Side(style="medium", color="555555")
    return Border(left=s, right=s, top=s, bottom=s)

def escribir_titulo(ws, texto, n_cols, fila, color_hex):
    ws.merge_cells(start_row=fila, start_column=1,
                   end_row=fila, end_column=n_cols)
    c = ws.cell(row=fila, column=1, value=f"  {texto}")
    c.fill      = PatternFill("solid", fgColor=color_hex)
    c.font      = Font(bold=True, color="FFFFFF", size=12)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[fila].height = 26

def escribir_headers(ws, headers, fila, color_hex):
    color_sub = color_hex  # misma familia, un tono más claro aplicado en datos
    for c_idx, h in enumerate(headers, 1):
        c = ws.cell(row=fila, column=c_idx, value=h)
        c.fill      = PatternFill("solid", fgColor=color_hex)
        c.font      = Font(bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = borde_medio()
    ws.row_dimensions[fila].height = 20

def escribir_fila(ws, datos, fila, color_par, es_par, formatos=None):
    fill = PatternFill("solid", fgColor="F2F2F2" if es_par else "FFFFFF")
    for c_idx, valor in enumerate(datos, 1):
        c = ws.cell(row=fila, column=c_idx, value=valor)
        c.fill      = fill
        c.font      = Font(size=10)
        c.border    = borde_fino()
        fmt = formatos.get(c_idx, None) if formatos else None
        if fmt:
            c.number_format = fmt
            c.alignment = Alignment(horizontal="right", vertical="center")
        elif isinstance(valor, (int, float)):
            c.alignment = Alignment(horizontal="right", vertical="center")
        else:
            c.alignment = Alignment(horizontal="left", vertical="center")

def escribir_subtotal(ws, datos, fila, n_cols, color_hex):
    for c_idx in range(1, n_cols + 1):
        valor = datos[c_idx - 1] if c_idx <= len(datos) else None
        c = ws.cell(row=fila, column=c_idx, value=valor)
        c.fill   = PatternFill("solid", fgColor=color_hex)
        c.font   = Font(bold=True, color="FFFFFF", size=10)
        c.border = borde_medio()
        c.alignment = Alignment(horizontal="right" if isinstance(valor, (int, float))
                                else "left", vertical="center")
    ws.row_dimensions[fila].height = 18

def ajustar_columnas(ws, df_ref, extra=4):
    for c_idx, col in enumerate(df_ref.columns, 1):
        max_len = max(len(str(col)),
                      df_ref[col].astype(str).str.len().max())
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max_len + extra, 40)


# ── Crear libro ───────────────────────────────
wb = Workbook()
wb.remove(wb.active)

# ─────────────────────────
#  Hoja 1: Ventas por Mercado
# ─────────────────────────
ws1 = wb.create_sheet("Ej1 - Ventas por Mercado")
ws1.sheet_view.showGridLines = False
COLOR_E1 = "1A5276"

escribir_titulo(ws1, "Ejercicio 1 · Ventas por Mercado", 4, 1, COLOR_E1)
escribir_titulo(ws1, "Ventas totales agrupadas por Mercado, ordenadas de mayor a menor",
                4, 2, "2980B9")
ws1.row_dimensions[2].height = 18
escribir_headers(ws1, ["#", "Mercado", "Ventas ($)", "% del Total"], 3, COLOR_E1)

t1_export = t1.reset_index()
t1_export.columns = ["#", "Mercado", "Ventas ($)", "% del Total"]
for i, row in t1_export.iterrows():
    escribir_fila(ws1, list(row), i + 4, COLOR_E1, i % 2 == 0,
                  formatos={3: '#,##0.00', 4: '0.00"%"'})

# Fila de total
fila_tot = len(t1_export) + 4
escribir_subtotal(ws1, ["", "TOTAL GLOBAL", round(ventas_total, 2), 100.00],
                  fila_tot, 4, COLOR_E1)
ws1.cell(fila_tot, 3).number_format = '#,##0.00'
ws1.cell(fila_tot, 4).number_format = '0.00"%"'

ws1.column_dimensions["A"].width = 5
ws1.column_dimensions["B"].width = 14
ws1.column_dimensions["C"].width = 18
ws1.column_dimensions["D"].width = 14
ws1.freeze_panes = "A4"

# ─────────────────────────
#  Hoja 2: Ventas por Año y Mes
# ─────────────────────────
ws2 = wb.create_sheet("Ej2 - Ventas por Año y Mes")
ws2.sheet_view.showGridLines = False
COLOR_E2 = "1E8449"

escribir_titulo(ws2, "Ejercicio 2 · Ventas por Año y Mes", 5, 1, COLOR_E2)
escribir_titulo(ws2, "Ventas mensuales de 2011 a 2014, con subtotales anuales",
                5, 2, "27AE60")
ws2.row_dimensions[2].height = 18
escribir_headers(ws2, ["Año", "Mes", "Nombre del Mes", "Ventas ($)",
                        "% del Año"], 3, COLOR_E2)

fila_act = 4
for año in [2011, 2012, 2013, 2014]:
    bloque = t2[t2["Año"] == año]
    subtotal_año = bloque["Ventas"].sum()
    for i, (_, row) in enumerate(bloque.iterrows()):
        pct_año = round(row["Ventas"] / subtotal_año * 100, 2)
        escribir_fila(ws2,
                      [int(row["Año"]), int(row["Mes"]),
                       row["Nombre_Mes"], round(row["Ventas"], 2), pct_año],
                      fila_act, COLOR_E2, i % 2 == 0,
                      formatos={4: '#,##0.00', 5: '0.00"%"'})
        fila_act += 1
    # Subtotal del año
    escribir_subtotal(ws2,
                      [str(año), "", "SUBTOTAL " + str(año),
                       round(subtotal_año, 2), 100.00],
                      fila_act, 5, COLOR_E2)
    ws2.cell(fila_act, 4).number_format = '#,##0.00'
    ws2.cell(fila_act, 5).number_format = '0.00"%"'
    fila_act += 1

# Total global
escribir_subtotal(ws2,
                  ["", "", "TOTAL GLOBAL", round(ventas_total, 2), ""],
                  fila_act, 5, "1A5276")
ws2.cell(fila_act, 4).number_format = '#,##0.00'

ws2.column_dimensions["A"].width = 8
ws2.column_dimensions["B"].width = 6
ws2.column_dimensions["C"].width = 17
ws2.column_dimensions["D"].width = 18
ws2.column_dimensions["E"].width = 13
ws2.freeze_panes = "A4"

# ─────────────────────────
#  Hoja 3: % Ventas por Método de Envío
# ─────────────────────────
ws3 = wb.create_sheet("Ej3 - Método de Envío")
ws3.sheet_view.showGridLines = False
COLOR_E3 = "6C3483"

escribir_titulo(ws3, "Ejercicio 3 · % de Ventas por Método de Envío", 4, 1, COLOR_E3)
escribir_titulo(ws3, "Participación porcentual de cada método de envío sobre las ventas totales",
                4, 2, "8E44AD")
ws3.row_dimensions[2].height = 18
escribir_headers(ws3, ["#", "Método de Envío", "Ventas ($)", "% del Total"], 3, COLOR_E3)

t3_export = t3.reset_index()
t3_export.columns = ["#", "Método de Envío", "Ventas ($)", "% del Total"]
for i, row in t3_export.iterrows():
    escribir_fila(ws3, list(row), i + 4, COLOR_E3, i % 2 == 0,
                  formatos={3: '#,##0.00', 4: '0.00"%"'})

fila_tot3 = len(t3_export) + 4
escribir_subtotal(ws3, ["", "TOTAL GLOBAL", round(ventas_total, 2), 100.00],
                  fila_tot3, 4, COLOR_E3)
ws3.cell(fila_tot3, 3).number_format = '#,##0.00'
ws3.cell(fila_tot3, 4).number_format = '0.00"%"'

ws3.column_dimensions["A"].width = 5
ws3.column_dimensions["B"].width = 20
ws3.column_dimensions["C"].width = 18
ws3.column_dimensions["D"].width = 14
ws3.freeze_panes = "A4"

# ─────────────────────────
#  Hoja 4: Top 3 Mercados por Prioridad
# ─────────────────────────
ws4 = wb.create_sheet("Ej4 - Top 3 por Prioridad")
ws4.sheet_view.showGridLines = False
COLOR_E4 = "784212"

escribir_titulo(ws4, "Ejercicio 4 · Top 3 Mercados con más Ventas por Prioridad de Envío",
                5, 1, COLOR_E4)
escribir_titulo(ws4, "Para cada nivel de prioridad, los 3 mercados con mayor volumen de ventas",
                5, 2, "E67E22")
ws4.row_dimensions[2].height = 18
escribir_headers(ws4,
                 ["Prioridad de Envío", "Puesto", "Mercado",
                  "Ventas ($)", "% de su Prioridad"], 3, COLOR_E4)

MEDALLAS = {1: "🥇 1°", 2: "🥈 2°", 3: "🥉 3°"}
COLORES_PRIORIDAD = {
    "Critical": "C0392B",
    "High":     "E67E22",
    "Medium":   "27AE60",
    "Low":      "2980B9",
}

fila_act4 = 4
for prioridad in ORDEN_PRIORIDAD:
    bloque = t4[t4["Prioridad de Envío"] == prioridad]
    total_prior = t4_full[t4_full["Prioridad Envio"] == prioridad]["Ventas"].sum()
    color_prior = COLORES_PRIORIDAD[prioridad]

    for i, (_, row) in enumerate(bloque.iterrows()):
        es_par = i % 2 == 0
        fill_c = PatternFill("solid", fgColor="F2F2F2" if es_par else "FFFFFF")
        datos = [
            prioridad if i == 0 else "",
            MEDALLAS.get(row["Puesto"], str(row["Puesto"])),
            row["Mercado"],
            round(row["Ventas"], 2),
            round(row["% de su Prioridad"], 2),
        ]
        for c_idx, valor in enumerate(datos, 1):
            c = ws4.cell(row=fila_act4, column=c_idx, value=valor)
            c.fill   = fill_c
            c.font   = Font(size=10)
            c.border = borde_fino()
            if c_idx in (4, 5):
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.number_format = '#,##0.00' if c_idx == 4 else '0.00"%"'
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")
            # Destacar la primera columna de cada prioridad
            if c_idx == 1 and i == 0:
                c.font = Font(bold=True, size=10, color=color_prior)
        fila_act4 += 1

    # Subtotal por prioridad
    escribir_subtotal(ws4,
                      [f"Total {prioridad}", "", f"{len(bloque)} mercados",
                       round(total_prior, 2), 100.00],
                      fila_act4, 5, color_prior)
    ws4.cell(fila_act4, 4).number_format = '#,##0.00'
    ws4.cell(fila_act4, 5).number_format = '0.00"%"'
    fila_act4 += 1

ws4.column_dimensions["A"].width = 22
ws4.column_dimensions["B"].width = 10
ws4.column_dimensions["C"].width = 12
ws4.column_dimensions["D"].width = 18
ws4.column_dimensions["E"].width = 20
ws4.freeze_panes = "A4"

# ── Guardar ───────────────────────────────────
wb.save(NOMBRE_ARCHIVO)
print(f"  Archivo guardado: {NOMBRE_ARCHIVO}")


# ── Resumen ───────────────────────────────────
print("\n" + "=" * 65)
print("  RESUMEN — Tablas generadas")
print("=" * 65)
print(f"  Ejercicio 1: {len(t1)} mercados analizados")
print(f"  Ejercicio 2: {len(t2)} registros (48 meses, 2011-2014)")
print(f"  Ejercicio 3: {len(t3)} métodos de envío analizados")
print(f"  Ejercicio 4: {len(t4)} registros (top 3 × 4 prioridades)")
print(f"\n  Ventas totales del dataset: ${ventas_total:,.2f}")
print("\n¡Paso 6 completado exitosamente!")
print("¡Actividad completada al 100%!")
